import csv
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

ARCHIVE_PATH = Path(__file__).parent.parent.joinpath('resources', 'mixed_files.zip')
FILES_DIR = Path(__file__).parent.parent.joinpath('data/test_files')


@pytest.fixture(scope='module', autouse=True)
def create_archive():
    with zipfile.ZipFile(ARCHIVE_PATH, 'w') as zip_file:
        for file in FILES_DIR.iterdir():
            zip_file.write(file, file.name)


def test_for_pdf_file(get_json_content):
    with zipfile.ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('content.pdf') as pdf_file:
            pdf_reader = PdfReader(pdf_file)
            pages_count = len(pdf_reader.pages)

            page = pdf_reader.pages[0]
            pdf_text = page.extract_text()
            pdf_clients_count = pdf_text.count('name')

            pdf_image_name = page.images[0].name

            assert pages_count == 1
            assert pdf_clients_count == len(get_json_content)
            assert pdf_image_name.endswith('.jpg')


def test_for_csv_file(get_json_content):
    with zipfile.ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('content.csv') as csv_file:
            csv_content = csv_file.read().decode()
            csv_reader = list(csv.reader(csv_content.splitlines()))
            keys_in_json = list(get_json_content[0].keys())
            headers_in_csv = csv_reader[0]
            first_user = csv_reader[1]

            assert headers_in_csv == keys_in_json
            assert first_user[0] == 'Anna'


def test_for_xlsx_file():
    with zipfile.ZipFile(ARCHIVE_PATH) as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_col=6, max_col=6, min_row=2, values_only=True):
                for age in row:
                    assert age >= 18
